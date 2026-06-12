#!/usr/bin/env python3
"""
Merge Tom Tom Sched export Type/Track (and Sub-type) into events_processed.tags.

The export xlsx uses Sched event IDs in column A (e.g. 107), while ``events_processed.id``
is usually ``event_records.id``. Rows are matched by:

1. ``external_link`` containing ``/event/{sched_id}/`` (Tom Tom Sched URLs), and
2. Title: export "Title" column vs ``events_processed.name`` (normalized).

Fallback: if ``CAST(events_processed.id AS TEXT)`` equals the export ID and the title matches.

Uses ``DATABASE_URL`` (default in ``src.app.models``). Run with ``--dry-run`` first.

Example::

  python scripts/merge_tomtom_sched_tags.py \\
    --xlsx scripts/tomtomfestival2026-event-export-2026-04-08-10-35-21.xlsx

  python scripts/merge_tomtom_sched_tags.py --xlsx path/to/export.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Optional

# Project root on path
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from sqlalchemy import text

from src.app.constants import TAG_MAX_COUNT, _normalize_tag_token
from src.app.models import connection_scope

# Sched listing export: row 7 = human headers; data starts row 9 (row 8 is REQUIRED legend).
_XLSX_HEADER_ROW = 7
_XLSX_DATA_START_ROW = 9
_COL_ID = 0
_COL_TITLE = 1
_COL_TYPE_TRACK = 8
_COL_SUBTYPE = 9

_SCHED_EVENT_ID_RE = re.compile(r"/event/(\d+)(?:/|$|[?#])", re.I)


def _norm_title(s: Optional[str]) -> str:
    if not s:
        return ""
    t = re.sub(r"\s+", " ", str(s)).strip().lower()
    return t


def _parse_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    try:
        ws = wb.active
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=_XLSX_DATA_START_ROW, values_only=True):
            raw_id = row[_COL_ID] if row else None
            if raw_id is None:
                continue
            sid = str(raw_id).strip()
            if not sid or not sid.isdigit():
                continue
            title = row[_COL_TITLE] if row else None
            type_track = row[_COL_TYPE_TRACK] if row else None
            subtype = row[_COL_SUBTYPE] if row else None
            rows.append(
                {
                    "sched_id": sid,
                    "title": (str(title).strip() if title is not None else "") or "",
                    "type_track": (str(type_track).strip() if type_track is not None else "") or None,
                    "subtype": (str(subtype).strip() if subtype is not None else "") or None,
                }
            )
        return rows
    finally:
        wb.close()


def _sched_id_from_link(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    m = _SCHED_EVENT_ID_RE.search(url)
    return m.group(1) if m else None


def _merge_tag_csv(existing: Optional[str], *extras: Optional[str]) -> Optional[str]:
    """Append normalized extras to existing comma-separated tags; dedupe (case-insensitive); cap count/length."""
    seen_lower: set[str] = set()
    out: list[str] = []
    if existing:
        for part in existing.split(","):
            t = _normalize_tag_token(part)
            if t:
                seen_lower.add(t.lower())
                out.append(t)
    for e in extras:
        t = _normalize_tag_token(e)
        if t and t.lower() not in seen_lower:
            seen_lower.add(t.lower())
            out.append(t)
    if len(out) > TAG_MAX_COUNT:
        out = out[:TAG_MAX_COUNT]
    if not out:
        return None
    return ", ".join(out)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--xlsx",
        type=Path,
        default=_ROOT / "scripts/tomtomfestival2026-event-export-2026-04-08-10-35-21.xlsx",
        help="Path to Sched export .xlsx",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Print planned updates without writing to the database",
    )
    args = p.parse_args()

    if not args.xlsx.is_file():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        return 1

    export_rows = _parse_xlsx(args.xlsx)
    by_sched: dict[str, dict[str, Any]] = {}
    for r in export_rows:
        by_sched[r["sched_id"]] = r

    select_sql = text(
        """
        SELECT id, name, tags, external_link
        FROM events_processed
        """
    )
    update_sql = text(
        """
        UPDATE events_processed
        SET tags = :tags
        WHERE id = :id
        """
    )

    matched = 0
    updated = 0
    skipped_no_export = 0
    skipped_title = 0
    n_db = 0

    with connection_scope() as session:
        db_rows = session.execute(select_sql).mappings().all()
        n_db = len(db_rows)

        for row in db_rows:
            eid = row["id"]
            name = row["name"]
            tags = row["tags"]
            ext = row["external_link"]

            sched_id = _sched_id_from_link(ext)
            export = None
            if sched_id and sched_id in by_sched:
                export = by_sched[sched_id]
            else:
                sid_str = str(eid).strip()
                if sid_str.isdigit() and sid_str in by_sched:
                    export = by_sched[sid_str]

            if not export:
                skipped_no_export += 1
                continue

            if _norm_title(name) != _norm_title(export["title"]):
                skipped_title += 1
                continue

            extras: list[Optional[str]] = []
            if export.get("type_track"):
                extras.append(export["type_track"])
            if export.get("subtype"):
                extras.append(export["subtype"])

            new_tags = _merge_tag_csv(tags, *extras)
            if new_tags == (tags or None):
                matched += 1
                continue

            matched += 1
            updated += 1
            if args.dry_run:
                nm = name or ""
                nprev = nm[:72] + ("…" if len(nm) > 72 else "")
                print(f"[dry-run] id={eid} sched_id={sched_id or eid} name={nprev!r}")
                print(f"         tags: {tags!r} -> {new_tags!r}")
            else:
                session.execute(update_sql, {"id": eid, "tags": new_tags})

        if not args.dry_run:
            session.commit()

    print(
        f"Done. export_rows={len(export_rows)} unique_sched_ids={len(by_sched)} "
        f"db_rows={n_db} matched_title={matched} updated={updated} "
        f"skipped_no_export_row={skipped_no_export} skipped_title_mismatch={skipped_title}"
    )
    if args.dry_run:
        print("(no database writes; omit --dry-run to apply)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
