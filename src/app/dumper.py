from abc import ABC, abstractmethod

from .utils import val
from .config import MISSING
from .constants import DB_FIELDS
from .repository import _normalize_date_for_key

from sqlalchemy.orm.session import Session

from .models import ProcessedURL, EventRecord



class BaseDumper(ABC):

    @abstractmethod
    def dump_events(self, events: list[dict]) -> None:
        pass

class ExcelDumper(BaseDumper):

    def __init__(self, filename: str, ws_title = "Events"):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        self._get_column_letter = get_column_letter
        self.filename = filename
        # setup excel workbook
        self.wb = Workbook()
        #setup current tab/ worksheet
        self.ws = self.wb.active
        self.ws.title = ws_title
        # set headers for the worksheet and append them to the sheet
        self.headers = [
            "Title", "StartDate", "EndDate",
            "StartTime", "EndTime", "DayOfWeek", "IsWeekend", "TimeOfDay",
            "EventCategory", "Audience", "LocationType",
            "Address", "Organizer", "Phone", "Website", "ImageURL",
            "Description", "Event Page", "Scraped At"
        ]
        self.ws.append(self.headers)

    def dump_events(self, events: list[dict]) -> None:
        for e in events:
            self.ws.append([
                val(e.get("title"), MISSING),
                val(e.get("start_date"), MISSING),
                val(e.get("end_date"), MISSING),
                val(e.get("start_time"), MISSING),
                val(e.get("end_time"), MISSING),
                val(e.get("day_of_week"), MISSING),
                val(e.get("is_weekend"), MISSING),
                val(e.get("time_of_day"), MISSING),
                val(e.get("event_category"), MISSING),
                val(e.get("audience"), MISSING),
                val(e.get("location_type"), MISSING),
                val(e.get("address"), MISSING),
                val(e.get("organizer"), MISSING),
                val(e.get("phone"), MISSING),
                val(e.get("website"), MISSING),
                val(e.get("image_url"), MISSING),
                val(e.get("description"), MISSING),
                val(e.get("event_link"), MISSING),
                val(e.get("scraped_at"), MISSING),
            ])
        # Auto-size column widths to fit content (max width 60).
        for col in range(1, len(self.headers) + 1):
            max_len = 0
            col_letter = self._get_column_letter(col)
            for cell in self.ws[col_letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            self.ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
        self.wb.save(self.filename)

class PostgreSQLDumper(BaseDumper):
    """Requires session= from session_scope() or get_session()."""

    def __init__(self, session: Session):
        if session is None:
            raise TypeError("PostgreSQLDumper requires session= (e.g. from session_scope()).")
        self.db = session
        self._own_session = False

    def close(self):
        if self._own_session:
            self.db.close()

    def is_processed(self, url: str) -> bool:
        return self.db.query(ProcessedURL).filter_by(url=url).first() is not None

    def mark_processed(self, url: str):
        self.db.add(ProcessedURL(url=url))

    def upsert_event(self, event: dict):
        event_link = event["event_link"]
        start_date_key = _normalize_date_for_key(event.get("start_date"))
        db_event = {k: v for k, v in event.items() if k in DB_FIELDS}
        if db_event.get("start_date") is not None and hasattr(db_event["start_date"], "isoformat"):
            db_event["start_date"] = _normalize_date_for_key(db_event["start_date"])

        existing = self.db.query(EventRecord).filter_by(
            event_link=event_link, start_date=start_date_key
        ).first()
        if existing:
            for k, v in db_event.items():
                setattr(existing, k, v)
        else:
            self.db.add(EventRecord(**db_event))

    def commit(self):
        self.db.commit()

    def rollback(self):
        self.db.rollback()

    def dump_events(self, events: list[dict]) -> None:
        try:
            for event in events:
                self.upsert_event(event)
            self.commit()
        except Exception:
            self.rollback()
            raise

