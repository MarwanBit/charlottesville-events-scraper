from datetime import datetime, timezone
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import json
from sqlalchemy.orm import sessionmaker
from models import engine, ProcessedURL, EventRecord

BASE_URL = "https://www.visitcharlottesville.org/events/"
HEADERS = {"User-Agent": "Mozilla/5.0"}
MISSING = "NULL"

MAX_PAGES = 30
STOP_AFTER_EMPTY_PAGES = 2


def clean_text(text):
    if not text:
        return ""
    return " ".join(text.split())


def val(x):
    if x is None:
        return MISSING
    s = str(x).strip()
    return s if s else MISSING


DB_FIELDS = {
    "event_link", "title",
    "start_date", "end_date", "start_time", "end_time",
    "day_of_week", "is_weekend", "time_of_day",
    "event_category", "audience", "location_type",
    "address", "organizer", "phone", "website", "image_url", "description",
    "scraped_at"
}

EVENT_TYPE_RULES = [
    ("Music", ["live music", "concert", "jazz", "rock", "band", "dj", "open mic", "karaoke"]),
    (
        "Food & Drink",
        ["wine", "beer", "brewery", "tasting", "cocktail", "dinner", "brunch", "restaurant", "food truck"]),
    ("Art", ["gallery", "exhibit", "exhibition", "art show", "museum", "artist talk"]),
    ("Theater", ["theatre", "theater", "play", "performance", "stage"]),
    ("Workshop/Class", ["workshop", "class", "lesson", "training", "bootcamp", "hands-on"]),
    ("Family/Kids", ["kids", "children", "family", "storytime", "toddler"]),
    ("Sports/Outdoor", ["run", "race", "hike", "yoga", "fitness", "outdoor", "park", "trail"]),
    ("Community/Networking", ["networking", "meetup", "community", "fundraiser", "volunteer", "charity"]),
    ("Education/Lecture", ["lecture", "talk", "seminar", "panel", "discussion", "university"]),
    ("Holiday/Seasonal", ["holiday", "christmas", "new year", "valentine", "halloween", "thanksgiving"]),
]

LOCATION_TYPE_RULES = [
    ("Winery/Brewery", ["winery", "vineyard", "brewery", "taproom"]),
    ("Restaurant/Bar", ["restaurant", "bar", "cafe", "coffee", "bistro", "pub"]),
    ("Theater/Venue", ["theatre", "theater", "auditorium", "venue"]),
    ("Museum/Gallery", ["museum", "gallery"]),
    ("University/School", ["university", "college", "school"]),
    ("Park/Outdoor", ["park", "garden", "trail", "outdoor"]),
    ("Hotel", ["hotel", "inn"]),
    ("Community Center", ["community center", "center", "library"]),
]


def _contains_any(text: str, keywords: list[str]) -> bool:
    t = (text or "").lower()
    return any(k in t for k in keywords)


def categorize_event(title: str, description: str, address: str, organizer: str) -> dict:
    blob = " ".join([title or "", description or "", address or "", organizer or ""]).lower()

    event_category = "Other"
    for cat, kws in EVENT_TYPE_RULES:
        if _contains_any(blob, kws):
            event_category = cat
            break

    if _contains_any(blob, ["kids", "children", "toddler", "family"]):
        audience = "Families"
    elif _contains_any(blob, ["21+", "adults only", "cocktail", "wine", "beer"]):
        audience = "Adults"
    else:
        audience = "All Ages"

    location_type = "Other"
    for loc, kws in LOCATION_TYPE_RULES:
        if _contains_any(blob, kws):
            location_type = loc
            break

    return {
        "event_category": event_category,
        "audience": audience,
        "location_type": location_type,
    }


def add_date_time_features(start_date: str, start_time: str) -> dict:
    day_of_week = ""
    is_weekend = ""
    time_of_day = ""

    try:
        if start_date:
            d = datetime.strptime(start_date, "%Y-%m-%d")
            day_of_week = d.strftime("%A")
            is_weekend = "TRUE" if day_of_week in ("Saturday", "Sunday") else "FALSE"
    except Exception:
        pass

    try:
        if start_time and re.match(r"^\d{2}:\d{2}$", start_time):
            hour = int(start_time.split(":")[0])
            if 5 <= hour < 12:
                time_of_day = "Morning"
            elif 12 <= hour < 17:
                time_of_day = "Afternoon"
            elif 17 <= hour < 21:
                time_of_day = "Evening"
            else:
                time_of_day = "Night"
    except Exception:
        pass

    return {
        "day_of_week": day_of_week,
        "is_weekend": is_weekend,
        "time_of_day": time_of_day,
    }


def get_image_url(soup: BeautifulSoup, base_url):
    div = soup.select_one(".background-image")
    if not div:
        return ""

    style = div.get("style", "")
    if "url(" in style:
        url = style.split("url(")[1].split(")")[0].strip("\"'")
        return urljoin(base_url, url)

    bgset = div.get("data-bgset", "")
    if bgset:
        return urljoin(base_url, bgset.split(",")[-1].split()[0])

    return ""


def extract_contact_info(soup):
    organizer = ""
    for li in soup.select("ul.detail__info li"):
        text = li.get_text(" ", strip=True)
        if text.lower().startswith("contact:"):
            organizer = text.split(":", 1)[1].strip()
    return organizer


def parse_subheading_to_dates_times(soup: BeautifulSoup):
    el = soup.select_one("p.page-title__subheading")
    if not el:
        return "", "", "", ""

    text = clean_text(el.get_text(" ", strip=True))
    time_match = re.search(r"(\d{1,2}:\d{2}\s*[AP]M)\s+to\s+(\d{1,2}:\d{2}\s*[AP]M)", text, re.IGNORECASE)

    start_time_24 = ""
    end_time_24 = ""
    if time_match:
        t1, t2 = time_match.groups()
        start_time_24 = datetime.strptime(t1.replace(" ", ""), "%I:%M%p").strftime("%H:%M")
        end_time_24 = datetime.strptime(t2.replace(" ", ""), "%I:%M%p").strftime("%H:%M")

    range_match = re.search(
        r"([A-Za-z]+)\s+(\d{1,2})\s+to\s+([A-Za-z]+)\s+(\d{1,2})",
        text,
        re.IGNORECASE
    )

    single_match = re.search(
        r"([A-Za-z]+)\s+(\d{1,2})(?!\s+to\s+[A-Za-z]+)",
        text,
        re.IGNORECASE
    )

    year = datetime.now().year

    if range_match:
        sm, sd, em, ed = range_match.groups()
        start_date_dt = datetime.strptime(f"{sm} {sd} {year}", "%B %d %Y")
        end_date_dt = datetime.strptime(f"{em} {ed} {year}", "%B %d %Y")

        if end_date_dt < start_date_dt:
            end_date_dt = end_date_dt.replace(year=year + 1)

        start_date = start_date_dt.strftime("%Y-%m-%d")
        end_date = end_date_dt.strftime("%Y-%m-%d")
        return start_date, end_date, start_time_24, end_time_24

    if single_match:
        m, d = single_match.groups()
        start_date_dt = datetime.strptime(f"{m} {d} {year}", "%B %d %Y")
        start_date = start_date_dt.strftime("%Y-%m-%d")
        return start_date, start_date, start_time_24, end_time_24

    return "", "", "", ""


def parse_jsonld_event(soup: BeautifulSoup):
    el = soup.select_one("script[type='application/ld+json']")
    if not el or not el.string:
        return {}

    try:
        data = json.loads(el.string)
        if isinstance(data, list):
            data = data[0] if data else {}
        if not isinstance(data, dict):
            return {}
        return data
    except Exception:
        return {}


def export_to_excel(events, filename="events.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"

    headers = [
        "Title", "StartDate", "EndDate",
        "StartTime", "EndTime", "DayOfWeek", "IsWeekend", "TimeOfDay",
        "EventCategory", "Audience", "LocationType",
        "Address", "Organizer", "Phone", "Website", "ImageURL",
        "Description", "Event Page", "Scraped At"
    ]

    ws.append(headers)

    for e in events:
        ws.append([
            val(e.get("title")),
            val(e.get("start_date")),
            val(e.get("end_date")),
            val(e.get("start_time")),
            val(e.get("end_time")),
            val(e.get("day_of_week")),
            val(e.get("is_weekend")),
            val(e.get("time_of_day")),
            val(e.get("event_category")),
            val(e.get("audience")),
            val(e.get("location_type")),
            val(e.get("address")),
            val(e.get("organizer")),
            val(e.get("phone")),
            val(e.get("website")),
            val(e.get("image_url")),
            val(e.get("description")),
            val(e.get("event_link")),
            val(e.get("scraped_at"))
        ])

    for col in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(filename)


def scrape_event_detail(session: requests.Session, event_url):
    resp = session.get(event_url, timeout=20)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    def pick_text(selectors):
        for sel in selectors:
            el = soup.select_one(sel)
            if el and el.get_text(strip=True):
                return clean_text(el.get_text(" ", strip=True))
        return ""

    description = pick_text([".text__text"])
    image_url = get_image_url(soup, event_url)
    organizer = extract_contact_info(soup)

    start_date = end_date = start_time = end_time = ""
    ld = parse_jsonld_event(soup)

    if ld:
        sd = ld.get("startDate", "") or ""
        ed = ld.get("endDate", "") or ""

        if sd:
            start_date = sd[:10] if len(sd) >= 10 else ""
            if "T" in sd:
                start_time = sd.split("T", 1)[1][:5]
        if ed:
            end_date = ed[:10] if len(ed) >= 10 else ""
            if "T" in ed:
                end_time = ed.split("T", 1)[1][:5]

        if not description and isinstance(ld.get("description"), str):
            description = clean_text(ld["description"])

        if not image_url:
            img = ld.get("image")
            if isinstance(img, str):
                image_url = urljoin(event_url, img)
            elif isinstance(img, list) and img and isinstance(img[0], str):
                image_url = urljoin(event_url, img[0])

    if not start_date or not end_date:
        sd2, ed2, st2, et2 = parse_subheading_to_dates_times(soup)
        start_date = start_date or sd2
        end_date = end_date or ed2
        start_time = start_time or st2
        end_time = end_time or et2

    return {
        "description": description,
        "image_url": image_url,
        "organizer": organizer,
        "start_date": start_date,
        "end_date": end_date,
        "start_time": start_time,
        "end_time": end_time,
    }


def main():
    session = requests.Session()
    session.headers.update(HEADERS)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    db = SessionLocal()

    page = 1
    all_events = []

    empty_pages = 0

    while True:
        url = BASE_URL if page == 1 else f"{BASE_URL}?page={page}"
        resp = session.get(url, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        cards = soup.select(".card")
        if not cards:
            break

        new_count = 0

        for card in cards:
            title_el = card.select_one(".card__heading a")
            date_el = card.select_one(".card__date-heading")
            address_el = card.select_one(".card__address")
            phone_el = card.select_one(".card__phone a")
            website_el = card.select_one(".card__website a")

            event_link = ""
            if title_el and title_el.get("href"):
                event_link = urljoin(BASE_URL, title_el["href"])

            title = clean_text(title_el.get_text()) if title_el else ""
            date = clean_text(date_el.get_text()) if date_el else ""
            address = clean_text(address_el.get_text()) if address_el else ""
            phone = clean_text(phone_el.get_text()) if phone_el else ""
            website = urljoin(event_link, website_el.get("href", "")) if website_el else ""

            if not title or not event_link:
                continue

            already_done = db.query(ProcessedURL).filter_by(url=event_link).first()
            if already_done:
                continue

            event = {
                "event_link": event_link,
                "title": title,
                "date": date,
                "address": address,
                "phone": phone,
                "website": website,
                "scraped_at": datetime.now(timezone.utc).isoformat()
            }

            try:
                details = scrape_event_detail(session, event_link)
                event.update(details)
                cats = categorize_event(
                    title=event.get("title", ""),
                    description=event.get("description", ""),
                    address=event.get("address", ""),
                    organizer=event.get("organizer", ""),
                )
                event.update(cats)

                dt_features = add_date_time_features(
                    start_date=event.get("start_date", ""),
                    start_time=event.get("start_time", ""),
                )
                event.update(dt_features)

                all_events.append(event)
                db_event = {k: v for k, v in event.items() if k in DB_FIELDS}

                existing = db.query(EventRecord).filter_by(event_link=event_link).first()
                if existing:
                    for k, v in db_event.items():
                        setattr(existing, k, v)
                else:
                    db.add(EventRecord(**db_event))

                db.add(ProcessedURL(url=event_link))  # ✅ add marker
                db.commit()

                new_count += 1

                print(f"\n📄 Scraping LISTING page {page} ...")
                print(f"  ▶ Event: {title}")
            except Exception as e:
                print("Detail scrape failed:", event_link, e)
                db.rollback()
                continue

        if new_count == 0:
            empty_pages += 1
        else:
            empty_pages = 0

        if empty_pages >= STOP_AFTER_EMPTY_PAGES:
            print(f"\n🛑 Stopping early: {empty_pages} pages in a row had 0 new events.")
            break

        # safety stop
        if page > MAX_PAGES:
            print(f"\n🛑 Safety stop: reached MAX_PAGES={MAX_PAGES}.")
            break

        page += 1

    export_to_excel(all_events, "../../visitcharlottesville_events.xlsx")
    db.close()
    print("\nDone ✅")
    print("Excel saved ✅ visitcharlottesville_events.xlsx")


if __name__ == "__main__":
    main()
