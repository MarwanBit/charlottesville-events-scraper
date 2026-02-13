from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from utils import val
from config import MISSING


def export_to_excel(events, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"

    headers = [
        "Title", "StartDate", "EndDate",
        "StartTime", "EndTime", "DayOfWeek", "IsWeekend", "TimeOfDay",
        "EventCategory", "Audience", "LocationType",
        "Address", "Organizer", "Phone", "Website", "ImageURL",
        "Description", "Event Page", "Scraped At"
    ]
    ws.append(headers)

    for e in events:
        ws.append([
            val(e.get("title"), MISSING),
            val(e.get("start_date"), MISSING),
            val(e.get("end_date"), MISSING),
            val(e.get("start_time"), MISSING),
            val(e.get("end_time"), MISSING),
            val(e.get("day_of_week"), MISSING),
            val(e.get("is_weekend"), MISSING),
            val(e.get("time_of_day"), MISSING),
            val(e.get("event_category"), MISSING),
            val(e.get("audience"), MISSING),
            val(e.get("location_type"), MISSING),
            val(e.get("address"), MISSING),
            val(e.get("organizer"), MISSING),
            val(e.get("phone"), MISSING),
            val(e.get("website"), MISSING),
            val(e.get("image_url"), MISSING),
            val(e.get("description"), MISSING),
            val(e.get("event_link"), MISSING),
            val(e.get("scraped_at"), MISSING),
        ])

    for col in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(filename)
