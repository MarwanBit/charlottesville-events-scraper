"""Unit tests for ExcelDumper: create worksheet, dump events, save, reopen, and verify content."""
from openpyxl import load_workbook

from src.app.dumper import ExcelDumper
from src.app.utils import val
from src.app.config import MISSING


def _expected_row(event: dict) -> list:
    """Build the same row the dumper writes for a given event dict."""
    return [
        val(event.get("title"), MISSING),
        val(event.get("start_date"), MISSING),
        val(event.get("end_date"), MISSING),
        val(event.get("start_time"), MISSING),
        val(event.get("end_time"), MISSING),
        val(event.get("day_of_week"), MISSING),
        val(event.get("is_weekend"), MISSING),
        val(event.get("time_of_day"), MISSING),
        val(event.get("event_category"), MISSING),
        val(event.get("audience"), MISSING),
        val(event.get("location_type"), MISSING),
        val(event.get("address"), MISSING),
        val(event.get("organizer"), MISSING),
        val(event.get("phone"), MISSING),
        val(event.get("website"), MISSING),
        val(event.get("image_url"), MISSING),
        val(event.get("description"), MISSING),
        val(event.get("event_link"), MISSING),
        val(event.get("scraped_at"), MISSING),
    ]


def test_excel_dumper_roundtrip(tmp_path):
    """Create a worksheet, dump events, save, reopen, and assert the events are present."""
    filepath = tmp_path / "test_events.xlsx"
    events = [
        {
            "title": "Test Event One",
            "start_date": "2025-03-15",
            "end_date": "2025-03-16",
            "event_category": "Music",
            "address": "123 Main St",
            "event_link": "https://example.com/event/1",
        },
        {
            "title": "Test Event Two",
            "start_date": "2025-04-01",
            "event_category": "Food & Drink",
            "organizer": "Local Org",
        },
    ]

    dumper = ExcelDumper(str(filepath), ws_title="Events")
    dumper.dump_events(events)

    wb = load_workbook(filepath, read_only=False)
    ws = wb.active

    assert ws.title == "Events"
    assert [c.value for c in ws[1]] == dumper.headers

    for i, event in enumerate(events):
        row_idx = i + 2
        row_cells = list(ws[row_idx])
        row_values = [c.value for c in row_cells]
        expected = _expected_row(event)
        assert row_values == expected, f"Row {row_idx} mismatch: got {row_values!r}"

    wb.close()


def test_excel_dumper_empty_events(tmp_path):
    """Dumping no events still creates a valid file with headers only."""
    filepath = tmp_path / "empty_events.xlsx"
    dumper = ExcelDumper(str(filepath))
    dumper.dump_events([])

    wb = load_workbook(filepath)
    ws = wb.active
    assert ws.title == "Events"
    assert [c.value for c in ws[1]] == dumper.headers
    assert ws.max_row == 1